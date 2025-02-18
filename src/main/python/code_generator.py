# from transformers import AutoTokenizer, AutoModelForSeq2SeqLM, T5ForConditionalGeneration
# from huggingface_hub import login
# import torch

# # Login to Hugging Face (replace with your token if needed)
# login(token='hf_ydtzoxrmoemdcYbXgosDkGy')

# # Initialize the CodeT5 model and tokenizer

# tokenizer = AutoTokenizer.from_pretrained("Salesforce/codet5-large")
# model = T5ForConditionalGeneration.from_pretrained("Salesforce/codet5-large")

# # Function to generate Angular code based on table data
# def generate_angular_code_from_table(table_data):
#     # Format the table data into a prompt
#     table_str = "\n".join([" | ".join(row) for row in table_data])
#     prompt = f"Generate an Angular component for the table with fields below:\n\n{table_str}\n\nAngular Component Code:"
#     print(prompt)
#     # Tokenize the input prompt
#     inputs_ids = tokenizer(prompt, return_tensors="pt").input_ids

#     generated_ids = model.generate(inputs_ids, max_length=2048, num_beams=5, early_stopping=True)


#     # Decode the generated output
#     generated_code = tokenizer.decode(generated_ids[0], skip_special_tokens=True)

#     # Extract generated code by removing the prompt
#     generated_code = generated_code.replace(prompt, '').strip()

#     # Return the generated Angular component code
#     return generated_code


import vertexai 
from vertexai.generative_models import GenerativeModel
import os
from openpyxl.utils import get_column_letter
import openpyxl
import torch
from transformers import AutoModelForCausalLM, AutoTokenizer
# Initialize the model and tokenizer
model_name = "nvidia/Llama-3.1-Nemotron-70B-Instruct-HF"
model = AutoModelForCausalLM.from_pretrained(model_name, torch_dtype=torch.bfloat16, device_map="auto")
tokenizer = AutoTokenizer.from_pretrained(model_name)


# # TODO(developer): Update and un-comment below line
# # PROJECT_ID = "your-project-id"
# vertexai.init(project="robotic-heaven-441813-p7", location="us-central1")
# model = GenerativeModel("gemini-1.5-pro-002")

# def generate_angular_code_from_table(table_data, prompt):
#     table_str = "\n".join([" | ".join(row) for row in table_data])
#     print("-------------------------",prompt,"------------------------")
#     print("-------------------------",table_str,"-------------------------------")
#     response = model.generate_content(
#         f"{prompt}:\n\n{table_str}"
#     )

#     # print(response.text)
#     return response.text      



def generate_angular_code_from_excel_files():
    # Iterate through each generated Excel file
    # table_index = 1
    # while os.path.exists(f"table_{table_index}.xlsx"):
    #     excel_filename = f"table_{table_index}.xlsx"
        
    #     # Read the Excel file and convert it to table data
    #     table_data = []
    #     wb = openpyxl.load_workbook(excel_filename)
    #     ws = wb.active

    #     for row in ws.iter_rows(values_only=True):
    #         table_data.append([str(cell) if cell is not None else "" for cell in row])

    #     # Prepare prompt for Vertex AI
    #     prompt = "Generate Angular form code based on the following table schema with validations"

    #     # Generate Angular code using Vertex AI
    #     response = model.generate_content(
    #         f"{prompt}:\n\n{table_data}"
    #     )

    #     angular_code = response.text
    #     output_filename = f"angular_code_{table_index}.ts"

    #     # Write the Angular code to a new file
    #     with open(output_filename, "w") as f:
    #         f.write(angular_code)

    #     print(f"Angular code generated and saved to: {output_filename}")
    #     table_index += 1
    if tokenizer.pad_token is None:
        tokenizer.pad_token = tokenizer.eos_token
    table_index = 1
    while os.path.exists(f"table_{table_index}.xlsx"):
        excel_filename = f"table_{table_index}.xlsx"
        
        # Read the Excel file and convert it to table data
        table_data = []
        wb = openpyxl.load_workbook(excel_filename)
        ws = wb.active

        for row in ws.iter_rows(values_only=True):
            table_data.append([str(cell) if cell is not None else "" for cell in row])

        # Convert table data to formatted string for the model input
        table_str = "\n".join([" | ".join(row) for row in table_data])

        # Prepare the prompt with the formatted table data
        prompt = f"Generate Angular code based on the following table data:\n\n{table_str}"

        # Use the prompt as a plain text input
        input_text = prompt

        # Tokenize the input text
        tokenized_input = tokenizer(
            input_text,
            return_tensors="pt",
            padding=True,
            truncation=True,
            max_length=1024
        )

        # Generate response from the model
        response_token_ids = model.generate(
            tokenized_input['input_ids'].to(model.device),
            attention_mask=tokenized_input['attention_mask'].to(model.device),
            max_new_tokens=4096,
            pad_token_id=tokenizer.eos_token_id
        )

        # Extract the generated text
        generated_tokens = response_token_ids[:, len(tokenized_input['input_ids'][0]):]
        generated_text = tokenizer.decode(generated_tokens[0], skip_special_tokens=True)

        # Save the generated Angular code to a file
        output_filename = f"angular_code_{table_index}.ts"
        with open(output_filename, "w") as f:
            f.write(generated_text)

        print(f"Angular code generated and saved to: {output_filename}")
        table_index += 1










