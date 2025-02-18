import openpyxl
from openpyxl.utils import get_column_letter
from bs4 import BeautifulSoup

# Function to extract all tables from HTML, create Excel files, and return the tables
def extract_all_tables_from_html_to_excel_and_return(html_content):
    # Parse the HTML content using BeautifulSoup
    soup = BeautifulSoup(html_content, 'html.parser')
    tables = soup.find_all('table')  # Find all table elements in the HTML
    table_index = 1  # To keep track of the table number for naming
    extracted_tables = []  # This will store the list of tables
    
    # Loop through each table
    for table in tables:
        # Extract rows from the table
        rows = table.find_all('tr')
        
        # Initialize table data to store the content of the current table
        table_data = []

        # Loop through each row in the table
        for row in rows:
            columns = row.find_all(['td', 'th'])  # Extract all cells (td and th)
            row_data = [column.get_text(strip=True) for column in columns]  # Extract text
            table_data.append(row_data)

        # Add the table's data to the list of extracted tables
        extracted_tables.append(table_data)
        
        # Create a new workbook and sheet for each table
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Table{table_index}"

        # Loop through each row in the table data to populate the Excel sheet
        for row_index, row in enumerate(table_data, 1):
            for col_index, cell_value in enumerate(row, 1):
                ws.cell(row=row_index, column=col_index, value=cell_value)

        # Adjust column widths based on content length
        for col in range(1, len(row_data) + 1):  # Loop through columns
            max_length = 0
            column = get_column_letter(col)
            for row in ws.iter_rows(min_col=col, max_col=col):
                for cell in row:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        # Save the Excel file with a unique name for each table
        wb.save(f"table_{table_index}.xlsx")
        print(f"Created excel with name: table_{table_index}.xlxs")
        table_index += 1
    
    # Return the list of extracted tables
    return extracted_tables